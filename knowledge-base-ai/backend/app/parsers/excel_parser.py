"""Excel document parser using openpyxl.

Extracts cell data from Excel workbooks (.xlsx), treating each worksheet as
a level-1 section and converting rows to human-readable text.

Supported format
----------------
``.xlsx`` only.  Legacy ``.xls`` (BIFF format) is **not** supported by
openpyxl.  Attempting to parse an ``.xls`` file raises ``ParserError`` with
a direction to convert the file to ``.xlsx`` first.

Row-to-text strategy
--------------------
Two heuristics are applied per sheet:

1. **Header row detection**: if the first non-empty row contains all string
   values, it is treated as a header row.  Subsequent rows are rendered as
   ``header: value`` pairs (one per cell), separated by newlines, to produce
   semantically rich text for embedding.

2. **No header row**: rows are rendered as tab-separated values — a neutral
   format that preserves structure without introducing false key-value
   semantics.

Empty cells are skipped to reduce noise.

Edge cases handled
------------------
- **Corrupted / invalid XLSX**: ``zipfile.BadZipFile``, ``KeyError``, and
  every other openpyxl exception are wrapped in ``ParserError``.
- **Password-protected workbook**: openpyxl cannot open encrypted workbooks;
  the ``InvalidFileException`` is caught and re-raised as ``ParserError``
  with a clear message.
- **Empty sheet**: skipped silently — produces no ``SectionContent`` entry and
  contributes no text to ``content``.
- **All-empty workbook**: returns ``ParsedDocument`` with empty ``content``
  (``is_empty == True``).
- **No page concept**: Excel has no pages.  ``pages`` always contains at most
  one synthetic entry; every ``SectionContent`` has ``page_number=None``.
"""
from __future__ import annotations

import os
import zipfile
from typing import Any

import structlog

from app.parsers.base import (
    DocumentParser,
    PageContent,
    ParsedDocument,
    ParserError,
    SectionContent,
)

log = structlog.get_logger(__name__)

# Maximum number of rows to process per sheet to guard against huge files.
# Beyond this limit the sheet is truncated and a WARNING is logged.
_MAX_ROWS_PER_SHEET = 10_000


def _cell_value(cell: Any) -> str:
    """Return a clean string representation of a cell value.

    ``None`` values become empty strings.  Numbers and dates are converted
    via ``str()`` for simplicity — good enough for embedding purposes.
    """
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _sheet_to_text(sheet: Any, file_path: str, logger: Any) -> str:
    """Convert all rows of an openpyxl worksheet to a readable text block.

    Parameters
    ----------
    sheet:
        An ``openpyxl.worksheet.worksheet.Worksheet`` instance.
    file_path:
        Used only for logging context.
    logger:
        A bound structlog logger.

    Returns
    -------
    str
        Multi-line text representation of the sheet, or an empty string
        if the sheet contains no non-empty cells.
    """
    rows: list[list[str]] = []
    total_rows = 0

    for row in sheet.iter_rows():
        if total_rows >= _MAX_ROWS_PER_SHEET:
            logger.warning(
                "excel_sheet_truncated",
                sheet=sheet.title,
                max_rows=_MAX_ROWS_PER_SHEET,
            )
            break
        cells = [_cell_value(c) for c in row]
        # Skip rows that are entirely empty.
        if any(cells):
            rows.append(cells)
        total_rows += 1

    if not rows:
        return ""

    # ── Header row heuristic ─────────────────────────────────────────────────
    # If every cell in the first row is a non-empty string that looks like a
    # label (no purely numeric content), treat it as the header.
    first_row = rows[0]
    has_header = bool(first_row) and all(
        v and not _looks_numeric(v) for v in first_row if v
    )

    lines: list[str] = []
    if has_header:
        headers = first_row
        for data_row in rows[1:]:
            pairs = [
                f"{headers[i]}: {data_row[i]}"
                for i in range(min(len(headers), len(data_row)))
                if i < len(data_row) and data_row[i]
            ]
            if pairs:
                lines.append("\n".join(pairs))
    else:
        for row in rows:
            line = "\t".join(v for v in row if v)
            if line:
                lines.append(line)

    return "\n\n".join(lines)


def _looks_numeric(value: str) -> bool:
    """Return True if *value* looks like a bare number (int or float)."""
    try:
        float(value)
        return True
    except ValueError:
        return False


class ExcelParser(DocumentParser):
    """Parse Excel workbooks (``.xlsx``) into a
    :class:`~app.parsers.base.ParsedDocument`.

    Each worksheet becomes a ``SectionContent`` with the sheet title as the
    heading (level 1).  Rows are converted to human-readable text using a
    header-detection heuristic.
    """

    @property
    def supported_extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    async def parse(self, file_path: str) -> ParsedDocument:
        """Extract text from an Excel workbook.

        Parameters
        ----------
        file_path:
            Absolute path to the ``.xlsx`` file on disk.

        Returns
        -------
        ParsedDocument
            Populated DTO with merged sheet content and per-sheet section
            metadata.

        Raises
        ------
        ParserError
            If the file is not found, is an unsupported format (``.xls``),
            is password-protected, is corrupt, or any other unrecoverable
            error occurs.
        """
        logger = log.bind(file_path=file_path)
        logger.info("excel_parse_start")

        if not os.path.exists(file_path):
            raise ParserError(
                file_path=file_path,
                reason="File not found",
            )

        # Reject legacy .xls files before attempting to open.
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xls":
            raise ParserError(
                file_path=file_path,
                reason=(
                    "Legacy .xls format is not supported. "
                    "Please convert the file to .xlsx (Excel 2007+) and re-ingest."
                ),
            )

        try:
            from openpyxl import load_workbook  # noqa: PLC0415
            workbook = load_workbook(
                file_path,
                read_only=True,   # memory-efficient; avoids loading styles
                data_only=True,   # return computed cell values, not formulas
            )
        except zipfile.BadZipFile as exc:
            raise ParserError(
                file_path=file_path,
                reason=f"File is not a valid XLSX (bad zip): {exc}",
            ) from exc
        except Exception as exc:
            # openpyxl raises InvalidFileException (subclass of Exception)
            # for password-protected workbooks and other open failures.
            exc_str = str(exc)
            if "encrypt" in exc_str.lower() or "password" in exc_str.lower():
                raise ParserError(
                    file_path=file_path,
                    reason=(
                        "Workbook appears to be password-protected and cannot "
                        "be opened without the password."
                    ),
                ) from exc
            raise ParserError(
                file_path=file_path,
                reason=f"Failed to open workbook: {exc}",
            ) from exc

        try:
            sections: list[SectionContent] = []
            content_parts: list[str] = []
            char_offset = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = _sheet_to_text(sheet, file_path, logger)

                if not sheet_text:
                    logger.info("excel_sheet_empty", sheet=sheet_name)
                    continue

                # Sheet title becomes the section heading text too.
                heading_line = sheet_name
                full_section_text = f"{heading_line}\n{sheet_text}"

                sections.append(
                    SectionContent(
                        heading=sheet_name,
                        level=1,
                        text=full_section_text,
                        page_number=None,
                        char_offset=char_offset,
                    )
                )
                content_parts.append(full_section_text)
                char_offset += len(full_section_text) + 1

            workbook.close()

            full_content = "\n\n".join(content_parts)
            pages: list[PageContent] = (
                [PageContent(page_number=1, text=full_content)]
                if full_content
                else []
            )

            file_name = os.path.basename(file_path)
            result = ParsedDocument(
                file_path=file_path,
                file_name=file_name,
                file_type=".xlsx",
                content=full_content,
                pages=pages,
                sections=sections,
            )

            logger.info(
                "excel_parse_complete",
                sheets_total=len(workbook.sheetnames),
                sheets_with_data=len(sections),
                content_length=len(full_content),
            )
            return result

        except ParserError:
            raise
        except Exception as exc:
            logger.error(
                "excel_parse_unexpected_error",
                error=str(exc),
                exc_info=True,
            )
            raise ParserError(
                file_path=file_path,
                reason=f"Unexpected error during Excel parsing: {exc}",
            ) from exc
